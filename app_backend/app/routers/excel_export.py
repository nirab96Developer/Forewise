"""
Excel Export Router — ייצוא נתונים ל-Excel
GET /api/v1/reports/export/excel?type=worklogs|invoices|projects|equipment
"""
import io
import urllib.parse
from datetime import date, datetime
from typing import Annotated, Literal

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import text

from app.core.database import get_db
from app.core.dependencies import get_current_active_user
from app.models.user import User

router = APIRouter(prefix="/reports/export", tags=["Excel Export"])

# ─── Style helpers ──────────────────────────────────────────────────────────

KKL_GREEN = "00994C"
HEADER_BG  = PatternFill("solid", fgColor=KKL_GREEN)
HEADER_FT  = Font(bold=True, color="FFFFFF", size=11)
ALT_ROW_BG = PatternFill("solid", fgColor="F0FFF7")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

def _style_header(ws, headers: list[str]) -> None:
    """Apply KKL green header row."""
    ws.sheet_view.rightToLeft = True
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FT
        cell.fill = HEADER_BG
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 22

def _style_data_rows(ws, num_rows: int, num_cols: int) -> None:
    for row in range(2, num_rows + 2):
        fill = ALT_ROW_BG if row % 2 == 0 else None
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            if fill:
                cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")

def _auto_width(ws, headers: list[str]) -> None:
    for col_idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(header)
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

def _workbook_to_response(wb: openpyxl.Workbook, filename: str) -> StreamingResponse:
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    # Use RFC 5987 encoding so Hebrew filenames work in all browsers
    encoded_name = urllib.parse.quote(filename, safe="")
    ascii_name = filename.encode("ascii", errors="ignore").decode() or "export.xlsx"
    content_disposition = f"attachment; filename=\"{ascii_name}\"; filename*=UTF-8''{encoded_name}"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": content_disposition},
    )


# ─── Worklogs ────────────────────────────────────────────────────────────────

def _export_worklogs(db: Session) -> openpyxl.Workbook:
    rows = db.execute(text("""
        SELECT
            wl.work_date,
            p.name         AS project_name,
            p.code         AS project_code,
            s.name         AS supplier_name,
            et.name        AS equipment_type,
            e.license_plate,
            wl.work_hours,
            wl.net_hours,
            wl.paid_hours,
            wl.hourly_rate_snapshot,
            wl.cost_before_vat,
            wl.cost_with_vat,
            wl.status,
            wl.is_overnight,
            wl.overnight_total
        FROM worklogs wl
        LEFT JOIN projects  p  ON p.id  = wl.project_id
        LEFT JOIN suppliers s  ON s.id  = wl.supplier_id
        LEFT JOIN equipment e  ON e.id  = wl.equipment_id
        LEFT JOIN equipment_types et ON et.id = e.equipment_type_id
        ORDER BY wl.work_date DESC
        LIMIT 5000
    """)).fetchall()

    headers = [
        "תאריך", "פרויקט", "קוד פרויקט", "ספק", "סוג ציוד",
        "לוחית רישוי", "שעות גולמי", "שעות נטו", "שעות לתשלום",
        "תעריף שעתי", "עלות ללא מע\"מ", "עלות עם מע\"מ",
        "סטטוס", "לינה", "סה\"כ לינה"
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "דיווחי שעות"
    _style_header(ws, headers)

    for r_idx, row in enumerate(rows, 2):
        ws.cell(r_idx, 1, str(row[0]) if row[0] else "")
        ws.cell(r_idx, 2, row[1] or "")
        ws.cell(r_idx, 3, row[2] or "")
        ws.cell(r_idx, 4, row[3] or "")
        ws.cell(r_idx, 5, row[4] or "")
        ws.cell(r_idx, 6, row[5] or "")
        ws.cell(r_idx, 7, float(row[6]) if row[6] else 0)
        ws.cell(r_idx, 8, float(row[7]) if row[7] else 0)
        ws.cell(r_idx, 9, float(row[8]) if row[8] else 0)
        ws.cell(r_idx, 10, float(row[9]) if row[9] else 0)
        ws.cell(r_idx, 11, float(row[10]) if row[10] else 0)
        ws.cell(r_idx, 12, float(row[11]) if row[11] else 0)
        ws.cell(r_idx, 13, row[12] or "")
        ws.cell(r_idx, 14, "כן" if row[13] else "לא")
        ws.cell(r_idx, 15, float(row[14]) if row[14] else 0)

    _style_data_rows(ws, len(rows), len(headers))
    _auto_width(ws, headers)
    return wb


# ─── Invoices ────────────────────────────────────────────────────────────────

def _export_invoices(db: Session) -> openpyxl.Workbook:
    rows = db.execute(text("""
        SELECT
            inv.invoice_number,
            s.name                              AS supplier_name,
            p.name                              AS project_name,
            p.code                              AS project_code,
            EXTRACT(MONTH FROM inv.issue_date)  AS month,
            EXTRACT(YEAR  FROM inv.issue_date)  AS year,
            inv.total_amount,
            inv.paid_amount,
            inv.status,
            inv.due_date,
            inv.created_at::date
        FROM invoices inv
        LEFT JOIN suppliers s ON s.id = inv.supplier_id
        LEFT JOIN projects  p ON p.id = inv.project_id
        ORDER BY inv.created_at DESC
        LIMIT 2000
    """)).fetchall()

    headers = [
        "מספר חשבונית", "ספק", "פרויקט", "קוד פרויקט",
        "חודש", "שנה", "סכום כולל", "שולם",
        "סטטוס", "תאריך פירעון", "תאריך יצירה"
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "חשבוניות"
    _style_header(ws, headers)

    for r_idx, row in enumerate(rows, 2):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(r_idx, c_idx)
            if val is None:
                cell.value = ""
            elif isinstance(val, (int, float)):
                cell.value = float(val)
            else:
                cell.value = str(val)

    _style_data_rows(ws, len(rows), len(headers))
    _auto_width(ws, headers)
    return wb


# ─── Projects ────────────────────────────────────────────────────────────────

def _export_projects(db: Session) -> openpyxl.Workbook:
    rows = db.execute(text("""
        SELECT
            p.code,
            p.name,
            r.name  AS region_name,
            a.name  AS area_name,
            p.status,
            p.start_date,
            p.end_date,
            b.total_amount,
            b.committed_amount,
            b.spent_amount,
            (b.total_amount - COALESCE(b.committed_amount,0) - COALESCE(b.spent_amount,0))
                    AS available,
            p.created_at::date
        FROM projects p
        LEFT JOIN regions r ON r.id = p.region_id
        LEFT JOIN areas   a ON a.id = p.area_id
        LEFT JOIN budgets b ON b.project_id = p.id AND b.is_active = true
        WHERE p.deleted_at IS NULL
        ORDER BY p.code
        LIMIT 5000
    """)).fetchall()

    headers = [
        "קוד", "שם פרויקט", "מרחב", "אזור", "סטטוס",
        "תאריך התחלה", "תאריך סיום",
        "תקציב כולל", "מוקפא", "נוצל", "זמין",
        "תאריך יצירה"
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "פרויקטים"
    _style_header(ws, headers)

    for r_idx, row in enumerate(rows, 2):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(r_idx, c_idx)
            if val is None:
                cell.value = ""
            elif hasattr(val, 'isoformat'):
                cell.value = str(val)
            elif isinstance(val, (int, float)):
                cell.value = float(val)
            else:
                cell.value = str(val)

    _style_data_rows(ws, len(rows), len(headers))
    _auto_width(ws, headers)
    return wb


# ─── Equipment ───────────────────────────────────────────────────────────────

def _export_equipment(db: Session) -> openpyxl.Workbook:
    rows = db.execute(text("""
        SELECT
            e.code,
            e.name,
            et.name  AS equipment_type,
            ec.name  AS category,
            s.name   AS supplier_name,
            se.license_plate,
            se.hourly_rate,
            et.overnight_rate,
            e.status,
            e.is_active
        FROM equipment e
        LEFT JOIN equipment_types      et ON et.id = e.equipment_type_id
        LEFT JOIN equipment_categories ec ON ec.id = et.category_id
        LEFT JOIN suppliers            s  ON s.id  = e.supplier_id
        LEFT JOIN supplier_equipment   se ON se.supplier_id = s.id
                                         AND se.is_active = true
        WHERE e.deleted_at IS NULL
        ORDER BY e.code
        LIMIT 5000
    """)).fetchall()

    headers = [
        "קוד", "שם", "סוג ציוד", "קטגוריה",
        "ספק", "לוחית רישוי",
        "תעריף שעתי", "תעריף לינה",
        "סטטוס", "פעיל"
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ציוד"
    _style_header(ws, headers)

    for r_idx, row in enumerate(rows, 2):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(r_idx, c_idx)
            if val is None:
                cell.value = ""
            elif isinstance(val, bool):
                cell.value = "כן" if val else "לא"
            elif isinstance(val, (int, float)):
                cell.value = float(val)
            else:
                cell.value = str(val)

    _style_data_rows(ws, len(rows), len(headers))
    _auto_width(ws, headers)
    return wb


# ─── Main endpoint ───────────────────────────────────────────────────────────

EXPORT_TYPES = Literal["worklogs", "invoices", "projects", "equipment"]

TYPE_LABELS = {
    "worklogs":  "worklogs",
    "invoices":  "invoices",
    "projects":  "projects",
    "equipment": "equipment",
}

TYPE_LABELS_HE = {
    "worklogs":  "דיווחי_שעות",
    "invoices":  "חשבוניות",
    "projects":  "פרויקטים",
    "equipment": "ציוד",
}

@router.get("/excel")
def export_excel(
    type: Annotated[EXPORT_TYPES, Query(description="סוג הייצוא")],
    db: Annotated[Session, Depends(get_db)],
    current_user: Annotated[User, Depends(get_current_active_user)],
):
    """
    ייצוא נתונים לקובץ Excel.
    type: worklogs | invoices | projects | equipment
    """
    today = date.today().strftime("%Y-%m-%d")
    label_he = TYPE_LABELS_HE.get(type, type)
    filename = f"forewise_{label_he}_{today}.xlsx"

    builders = {
        "worklogs":  _export_worklogs,
        "invoices":  _export_invoices,
        "projects":  _export_projects,
        "equipment": _export_equipment,
    }

    wb = builders[type](db)
    return _workbook_to_response(wb, filename)
